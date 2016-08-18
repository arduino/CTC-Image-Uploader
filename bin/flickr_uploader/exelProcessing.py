#
# Process the excel sheets for youtube videos and github code
# Retrive short url.
#
# state codes:
# 0: not shortened
# 1: shortened
#
import openpyxl
from mainDB import CTCPhotoDB
from CreateShortLinks import getShortURL
from ThreadWork import MultiThreadWork, MultiTaskSingleThreadWork
from configs import extras_def

photoDB=CTCPhotoDB()

tw=MultiThreadWork(3)
mainTW=MultiTaskSingleThreadWork()


#
# Util function for getting all the type names
#
#
def getTypeCodes():
	return [one["type_code"] for one in extras_def]
#
# Util function for getting the type name of an
# extra record
#
#
def getFullType(typeCode):
	return filter(lambda x: x["type_code"]==typeCode,extras_def)[0]["full_type"]

#
# Util function for getting the in-set order 
# by reading the last digits of a shortCode
#
#
def getOrderInSet(shortCode):
	identifier=shortCode.split("-")[-2:]
	if identifier[0].isdigit():
		return int(identifier[1])-1
	else:
		return 0

#
# Check if every field in toCompare object matches 
# the one in base
#
#
def recNotMatch(toCompare, base):
	for key in toCompare:
		if toCompare[key]!=base[key]:
			print type(toCompare[key]),type(base[key])
			return True
	return False


#
# get the active sheet of a openpyxl object from a file 
# location string
#
#
def getSheet(workbookLocation, sheetName=""):
	wb = openpyxl.load_workbook(workbookLocation)
	if not sheetName:
		sheet=wb.active
	else:
		sheet=wb[sheetName]
	return sheet

#
# Depending on if a record exists/is identical to the new one,
# make a new record/update the old record/do nothing
#
#
def addNewOrUpdateOld(newRec):
	shortCode=newRec["short_code"]
	oldRec=photoDB.getExtraByShortCode(shortCode)
	if oldRec==None:
		newRec["short_code"]=shortCode
		photoDB.addExtra(newRec)
		message="New record"
	elif recNotMatch(newRec,oldRec):
		newRec["state"]=0
		del newRec["short_code"]
		for key in newRec:
			if not (type(newRec[key]) is int):
				newRec[key]="'{}'".format(newRec[key])
		photoDB.modifyExtraByShortCode(shortCode, **newRec)
		message="Modifying"
	else:
		return
	print message, newRec["name"], shortCode



#
# Read data from the exel sheet and 
# save it in the database
#
#
def processSheet(sheet, nameCol, linkCol, shortCodeCol, rangeStart, type, shortCodePrefix):
	for row in range(rangeStart,sheet.max_row+1):
		name=unicode(sheet.cell(row=row,column=nameCol).value)
		link=sheet.cell(row=row,column=linkCol).value
		shortCode=sheet.cell(row=row,column=shortCodeCol).value
		if shortCode!=None and link!=None:
			orderInSet=getOrderInSet(shortCode)
			shortCode="ctc-"+shortCodePrefix+"-"+shortCode
			newRec={"name":name, "short_code":shortCode, "hosted_url":link, "type":type, "order_in_set":orderInSet}
			addNewOrUpdateOld(newRec)
	photoDB.commit()

#
# get a list of parameters for processSheet function, from an
# extras_def record
#
#
def getProcessSheetParams(definition):
	one=definition

	bookName=one["book_name"]
	sheetName=one.get("sheet_name","")
	sheet=getSheet(bookName,sheetName)

	typeCode=one["type_code"]
	shortCodePrefix=one.get("short_code_prefix",one["type_code"])

	sheetSettings=one["sheet_settings"]
	args=[sheet]
	args.extend(sheetSettings)
	args.extend([typeCode,shortCodePrefix])

	return args

#
# process all the sheets as defined in extras_def of configs file
#
#
def processAllSheets():
	for one in extras_def:
		print "Processing sheet: ", one["name"]
		params=getProcessSheetParams(one)
		processSheet(*params)
		print "-------------------------"

#
# Get all records from extras table that hasn't been shortened
#
#
def getUnshortenedExtras():
	cmd="""
	SELECT * FROM extras
	WHERE state == 0
	"""
	res=photoDB.makeQuery(cmd)[0].fetchall()
	return res




#
# MultiThreadWork task for getting the short URL and regist a
# database task
#
#
def shortenExtra(workerIndex, rec):
	res=getShortURL({ \
		"hosted_url":rec["hosted_url"], \
		"keyword":rec["short_code"], \
		"title":getFullType(rec["type"])+" "+rec["name"] \
		})
	if res:
		mainTW.addTask("saveState",rec["short_code"])

#
# MultiTaskSingleThreadWork task for changing an extra record
# as shortened
#
#
def saveState(task):
	photoDB.modifyExtraState(task,1)

#
# The main procedure for getting extras shortlinks 
#
#
def getShortURLForExtras():
	tw.setActualTask(shortenExtra)

	mainTW.addActualTaskType("saveState",saveState)

	unshortenedExtras=getUnshortenedExtras()
	for one in unshortenedExtras:
		print "Requesting shortLink",one["short_code"]
		tw.addTask(one)

	while tw.processing():
		mainTW.process()
	tw.join()
	mainTW.process()


#
# Output the extra sheet into a xml document
#
#
def outputXML():
	outputFile=open("testExtras.txt","w")
	
	output="<data>\n"
	for typeCode in getTypeCodes():
		output=output+"<category type='"+getFullType(typeCode)+"'>\n"
		recs=photoDB.getRecByField("extras","type","'"+typeCode+"'").fetchall()
		for rec in recs:
			output=output+"<rec>\n<title>{}</title>\n<link>{}</link>\n</rec>\n".format(rec["name"],"http://verkstad.cc/urler/"+rec["short_code"])
		output=output+"</category>\n"
	output=output+"</data>\n"
	
	outputFile.write(output)
	outputFile.close()



#
# Main procedure
#
#
def exelProcessing():
	print "Processing all extras sheets"
	processAllSheets()

	print "Retriving extras shortLinks"
	getShortURLForExtras()

	print "Outputing xml"
	outputXML()


if __name__=="__main__":
	for one in ["g","y","frz","ico","dld","fail"]:
		try:
			print getFullType(one)
		except Exception as e:
			print e

	print getTypeCodes()

	for one in extras_def:
		print getProcessSheetParams(one)

	processAllSheets()

